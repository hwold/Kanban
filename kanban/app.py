from flask import Flask, jsonify, request,render_template
from openpyxl import load_workbook,Workbook
import os
from datetime import datetime
import time
app = Flask(__name__)



# Caminho do arquivo Excel
excel_file = '.\kanban_data.xlsx'
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tasks'
    # Adicionando cabeçalhos
    ws.append(['id', 'titulo', 'descricao', 'responsavel','prazo','status','progresso','dataCriacao','dataConclusao','bullet'])
    wb.save(excel_file)
    print(f"Arquivo {excel_file} criado com sucesso.")
else:
    print(f"O arquivo {excel_file} já existe.")


def carregar_atividades():
   workbook = load_workbook(excel_file)
   sheet = workbook.active
   atividades = []
   for row in sheet.iter_rows(min_row=2, values_only=True):
       atividades.append({
           'id': row[0],
           'titulo': row[1],
           'descricao': row[2],
           'responsavel': row[3],
           'prazo': row[4],
           'status': row[5],
           'progresso': row[6],
           'dataCriacao': row[7],
           'dataConclusao': row[8],
           'bullet': row[9].split(';') if row[9] else []  # Carrega os bullet points
       })
   atividades_filtradas = []
   dias_limite = 4  # Defina o número de dias após os quais deseja ocultar atividades concluídas
   for atividade in atividades:
       if atividade['status'] == 'done' and atividade['dataConclusao']:
           # Converte a data de conclusão para um objeto datetime
           data_conclusao = datetime.strptime(atividade['dataConclusao'], '%Y-%m-%d')
           dias_passados = (datetime.now() - data_conclusao).days
           # Se a atividade foi concluída há mais do que o limite de dias, não adiciona à lista
           if dias_passados > dias_limite:
               continue  # Pula essa atividade
       # Se não estiver concluída há mais de "x" dias, ou se ainda estiver em andamento, adiciona à lista
       atividades_filtradas.append(atividade)
   return atividades_filtradas



# Salvar atividades no Excel
def salvar_atividades(atividades):
    time.sleep(1)
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Limpar a planilha antes de salvar os novos dados
    sheet.delete_rows(2, sheet.max_row)

    # Inserir os dados novamente
    for i, atividade in enumerate(atividades, start=2):
        atividade['bullet'] = [bullet for bullet in atividade['bullet'] if bullet.strip()]
        sheet[f'A{i}'] = atividade['id']
        sheet[f'B{i}'] = atividade['titulo']
        sheet[f'C{i}'] = atividade['descricao']
        sheet[f'D{i}'] = atividade['responsavel']
        sheet[f'E{i}'] = atividade['prazo']
        sheet[f'F{i}'] = atividade['status']
        sheet[f'G{i}'] = atividade['progresso']
        sheet[f'H{i}'] = atividade['dataCriacao']
        sheet[f'I{i}'] = atividade['dataConclusao']
        sheet[f'J{i}'] = ';'.join(atividade['bullet'])

    workbook.save(excel_file)

# Rota principal para carregar a página Kanban
@app.route('/')
def index():
    return render_template('index.html')

# Endpoint para carregar as atividades do Excel
@app.route('/atividades', methods=['GET'])
def get_atividades():
    atividades = carregar_atividades()
    return jsonify(atividades)

# Endpoint para salvar as atividades no Excel
@app.route('/atividades', methods=['POST'])
def post_atividades():
    atividades = request.json
    salvar_atividades(atividades)
    return jsonify({'message': 'Atividades salvas com sucesso!'})


if __name__ == '__main__':
    app.run(debug=True)

    

